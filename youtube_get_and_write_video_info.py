from googleapiclient.discovery import build
from openpyxl import load_workbook
from tqdm import tqdm
import pandas as pd
import re
import json


def GetYoutubeVideoInfo(youtube, id_list, start_index):
    response = youtube.videos().list(
        part="snippet,statistics,contentDetails",
        id=','.join(id_list[start_index:start_index + 50])
    ).execute()
    return response


def GetVideoIdList(urls_file_path):
    content = ""
    youtube_pattern = r"https?://www\.youtube\.com/watch\?v=([\w-]+)"

    with open(urls_file_path, 'r') as f:
        content = f.read()
    return re.findall(youtube_pattern, content)


def GetVideoDataObject(video_info, video_id):
    data = {"VideoId": "","Title": "","PublishedAt": "","ChannelTitle": "","ChannelId": "","Views": "","Likes": "","Comments": "","Duration": ""}
    data["VideoId"] = video_id
    data["Title"] = (video_info["snippet"]["title"] if "title" in video_info["snippet"] else "")
    data["PublishedAt"] = (video_info["snippet"]["publishedAt"] if "publishedAt" in video_info["snippet"] else "")
    data["ChannelTitle"] = (video_info["snippet"]["channelTitle"] if "channelTitle" in video_info["snippet"] else "")
    data["ChannelId"] = (video_info["snippet"]["channelId"] if "channelId" in video_info["snippet"] else "")
    data["Views"] = (video_info["statistics"]["viewCount"] if "viewCount" in video_info["statistics"] else "")
    data["Likes"] = (video_info["statistics"]["likeCount"] if "likeCount" in video_info["statistics"] else "")
    data["Comments"] = (video_info["statistics"]["commentCount"] if "commentCount" in video_info["statistics"] else "")
    data["Duration"] = (video_info["contentDetails"]["duration"] if "duration" in video_info["contentDetails"] else "")
    return data


def AppendToExcelSheet(entries, excel_path):
    df = pd.DataFrame(entries)

    book = load_workbook(excel_path)
    last_row = book["Sheet1"].max_row

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=last_row)


def main():
    json_data = None
    with open('paths.json', 'r') as j:
        json_data = json.load(j)
    
    API_KEY = json_data["API_KEY"]
    FILE_PATH = json_data["FILE_PATH"]
    URLS_PATH = json_data["URLS_PATH"]

    video_id_list = GetVideoIdList(URLS_PATH)
    youtube = build("youtube", "v3", developerKey=API_KEY)
    data_entries = []

    # getting response for only 50 ids at a time
    for index in tqdm(range(0, len(video_id_list), 50), desc="Processing"):
        try:
            response = GetYoutubeVideoInfo(youtube, video_id_list, index)
            video_info_list = response["items"]
        except Exception as e:
            print(f"Error occurred during API call -- index: {index} -- {e}")
            continue
        
        id_index = index
        for video_info in video_info_list:
            try:
                data = GetVideoDataObject(video_info, video_id_list[id_index])
                data_entries.append(data)
            except:
                print(f"Error in data processing for item : {video_info}")
            id_index += 1

    AppendToExcelSheet(data_entries, FILE_PATH)
    print("Process complete.")


if __name__ == '__main__':
    main()